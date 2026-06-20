import os
from datetime import datetime
from PyQt5.QtWidgets import QFileDialog
from app.db.dao import CostItemDAO, EventDAO, DocumentDAO


def export_summary(event_id, version_id=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    event = EventDAO.get_by_id(event_id)
    if not event:
        return None

    default_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '..', 'data', 'exports')
    default_dir = os.path.abspath(default_dir)
    if not os.path.exists(default_dir):
        os.makedirs(default_dir)
    default_name = f"索赔汇总表_{event.get('contract_section','')}_{event.get('start_date','')}.xlsx".replace('/', '-')
    path, _ = QFileDialog.getSaveFileName(None, '导出汇总表', os.path.join(default_dir, default_name), 'Excel 文件 (*.xlsx)')
    if not path:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = '索赔汇总'

    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    title_font = Font(name='微软雅黑', size=16, bold=True)
    header_font = Font(name='微软雅黑', size=11, bold=True)
    normal_font = Font(name='微软雅黑', size=10)
    header_fill = PatternFill('solid', fgColor='DCE6F1')
    total_fill = PatternFill('solid', fgColor='FCE4D6')

    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = '停窝工索赔费用汇总表'
    c.font = title_font
    c.alignment = center

    row = 3
    info = [
        ('事件编号', event.get('id', '')),
        ('事件类型', event.get('event_type', '')),
        ('合同段', event.get('contract_section', '')),
        ('影响部位', event.get('affected_area', '')),
        ('起止时间', f"{event.get('start_date', '')} 至 {event.get('end_date', '（进行中）')}"),
        ('责任方', event.get('responsible_party', '')),
        ('监理通知', event.get('supervision_notice_no', '')),
        ('业主指令', event.get('owner_order_no', '')),
    ]
    for k, v in info:
        ws.cell(row=row, column=1, value=k).font = header_font
        ws.cell(row=row, column=1).alignment = left
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        c = ws.cell(row=row, column=2, value=v)
        c.font = normal_font
        c.alignment = left
        row += 1

    row += 1

    ver_name = ''
    if version_id:
        from app.db.dao import CostVersionDAO
        for v in CostVersionDAO.get_versions(event_id):
            if v['id'] == version_id:
                ver_name = v.get('version_name', '')
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                c = ws.cell(row=row, column=1, value=f'测算版本：{ver_name}' + ('（当前版本）' if v.get('is_current') else ''))
                c.font = header_font
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = border
                row += 1
                break

    items_by_cat, totals, grand = CostItemDAO.get_summary(event_id, version_id)
    for cat in CostItemDAO.CATEGORIES:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=f'【{cat}】')
        c.font = header_font
        c.fill = header_fill
        c.alignment = left
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

        headers = ['序号', '项目名称', '单价(元)', '数量', '单位', '金额(元)', '备注']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.alignment = center
            c.border = border
        row += 1

        cat_items = items_by_cat.get(cat, [])
        if not cat_items:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            c = ws.cell(row=row, column=1, value='（无）')
            c.font = normal_font
            c.alignment = center
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = border
            row += 1
        for i, it in enumerate(cat_items, 1):
            values = [i, it['item_name'], it['unit_price'], it['quantity'],
                      it.get('unit', ''), it['amount'], it.get('remark', '')]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = normal_font
                c.alignment = center if col in (1, 3, 4, 5, 6) else left
                c.border = border
                if col in (3, 6):
                    c.number_format = '0.00'
            row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=f'{cat} 小计')
        c.font = header_font
        c.fill = total_fill
        c.alignment = Alignment(horizontal='right', vertical='center')
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).fill = total_fill
        c6 = ws.cell(row=row, column=6, value=totals[cat])
        c6.font = header_font
        c6.fill = total_fill
        c6.alignment = center
        c6.number_format = '0.00'
        c6.border = border
        ws.cell(row=row, column=7).fill = total_fill
        ws.cell(row=row, column=7).border = border
        row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value='索赔金额合计（大写）：' + CostCalcWidget_num2chinese(grand))
    c.font = Font(name='微软雅黑', size=11, bold=True)
    c.alignment = left
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value=f'索赔金额合计（小写）：￥{grand:,.2f}')
    c.font = Font(name='微软雅黑', size=14, bold=True, color='C00000')
    c.alignment = left
    row += 2

    doc_status_title = '支撑材料档案归档情况'
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value=doc_status_title)
    c.font = header_font
    c.fill = header_fill
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).fill = header_fill
    row += 1

    for t in DocumentDAO.TYPES:
        docs = DocumentDAO.get_by_event_and_type(event_id, t)
        if docs:
            status_text = f'✔ {t}：已归档 {len(docs)} 份'
            font_color = '008000'
        else:
            status_text = f'✘ {t}：未归档（待补）'
            font_color = 'C00000'
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=status_text)
        c.font = Font(name='微软雅黑', size=10, color=font_color)
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

        for d in docs:
            doc_info = f"   编号：{d.get('doc_no', '')}    文件：{d.get('file_path', '')}    备注：{d.get('remark', '')}"
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            c = ws.cell(row=row, column=1, value=doc_info)
            c.font = Font(name='微软雅黑', size=9, color='555555')
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = border
            row += 1

    missing = EventDAO.get_missing_docs(event_id)
    if missing:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value='⚠ 待补材料：' + '、'.join([m for m in missing if m not in DocumentDAO.TYPES]))
        c.font = Font(name='微软雅黑', size=10, color='C00000')
        row += 1

    row += 1

    row += 1
    sign = ['商务经理签字：______________', '项目经理签字：______________', f"日期：{datetime.now().strftime('%Y年%m月%d日')}"]
    for s in sign:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value=s).font = normal_font
        row += 1

    widths = [8, 28, 12, 10, 8, 14, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(path)
    return path


def export_monthly_archive(year_month=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from collections import defaultdict

    events = EventDAO.get_all()
    if not events:
        return None

    by_month = defaultdict(list)
    for e in events:
        sd = e.get('start_date', '')
        if not sd or len(sd) < 7:
            continue
        ym = sd[:7]
        if year_month and ym != year_month:
            continue
        by_month[ym].append(e)

    if not by_month:
        return None

    default_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '..', 'data', 'exports')
    default_dir = os.path.abspath(default_dir)
    if not os.path.exists(default_dir):
        os.makedirs(default_dir)
    ym_str = year_month if year_month else '全部'
    default_name = f"月度资料归档清单_{ym_str}.xlsx".replace('/', '-')
    path, _ = QFileDialog.getSaveFileName(None, '导出月度归档清单', os.path.join(default_dir, default_name), 'Excel 文件 (*.xlsx)')
    if not path:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = '资料归档清单'

    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    title_font = Font(name='微软雅黑', size=14, bold=True)
    header_font = Font(name='微软雅黑', size=10, bold=True)
    normal_font = Font(name='微软雅黑', size=10)
    header_fill = PatternFill('solid', fgColor='DCE6F1')
    ok_fill = PatternFill('solid', fgColor='E2EFDA')
    miss_fill = PatternFill('solid', fgColor='FCE4D6')
    month_fill = PatternFill('solid', fgColor='FFF2CC')

    doc_types = ['现场签证单', '复工令', '监理通知', '业主指令', '现场照片', '机械清单', '劳务记录']

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    c = ws['A1']
    c.value = f'停窝工索赔资料归档清单（{ym_str}）'
    c.font = title_font
    c.alignment = center

    headers = ['序号', '月份', '事件类型', '开始日期', '结束日期', '合同段', '影响部位'] + doc_types + ['状态汇总', '跟进状态', '备注']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.alignment = center
        c.border = border
        c.fill = header_fill

    row = 4
    idx = 1
    follow_status_map = {
        'pending_commercial': '待商务补',
        'pending_site': '待现场补',
        'reminded': '已催办',
        'closed': '已闭合',
        None: '未设置',
        '': '未设置'
    }

    for ym in sorted(by_month.keys(), reverse=True):
        month_events = sorted(by_month[ym], key=lambda x: x.get('start_date', ''))
        month_total = len(month_events)
        month_complete = 0

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        c = ws.cell(row=row, column=1, value=f'━━━ {ym} 月份（共 {month_total} 个事件）━━━')
        c.font = header_font
        c.fill = month_fill
        c.alignment = left
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).fill = month_fill
        row += 1

        for e in month_events:
            eid = e['id']
            missing = EventDAO.get_missing_docs(eid)
            docs = DocumentDAO.get_by_event(eid)
            photos = __import__('app.db.dao', fromlist=['PhotoDAO']).PhotoDAO.get_by_event(eid)
            machines = __import__('app.db.dao', fromlist=['MachineryDAO']).MachineryDAO.get_by_event(eid)
            labors = __import__('app.db.dao', fromlist=['LaborDAO']).LaborDAO.get_by_event(eid)

            doc_status = {}
            for t in DocumentDAO.TYPES:
                doc_status[t] = any(d['doc_type'] == t for d in docs)
            doc_status['现场照片'] = len(photos) > 0
            doc_status['机械清单'] = len(machines) > 0
            doc_status['劳务记录'] = len(labors) > 0

            all_ok = all(doc_status.values())
            if all_ok:
                month_complete += 1

            values = [
                idx,
                ym,
                e.get('event_type', ''),
                e.get('start_date', ''),
                e.get('end_date', ''),
                e.get('contract_section', ''),
                e.get('affected_area', '')
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = normal_font
                c.alignment = center if col in (1, 2, 3, 4, 5) else left
                c.border = border

            status_col = 8
            for t in doc_types:
                ok = doc_status[t]
                c = ws.cell(row=row, column=status_col, value='✓ 齐全' if ok else '✘ 待补')
                c.font = Font(name='微软雅黑', size=10, bold=True, color='008000' if ok else 'C00000')
                c.alignment = center
                c.border = border
                c.fill = ok_fill if ok else miss_fill
                status_col += 1

            if all_ok:
                summary = '✔ 资料齐全'
                sum_color = '008000'
                sum_fill = ok_fill
            else:
                summary = f'✘ 缺：{"、".join(missing[:3])}' + ('…' if len(missing) > 3 else '')
                sum_color = 'C00000'
                sum_fill = miss_fill
            c = ws.cell(row=row, column=15, value=summary)
            c.font = Font(name='微软雅黑', size=10, bold=True, color=sum_color)
            c.alignment = left
            c.border = border
            c.fill = sum_fill

            fs = e.get('follow_status', '')
            c = ws.cell(row=row, column=16, value=follow_status_map.get(fs, '未设置'))
            c.font = normal_font
            c.alignment = center
            c.border = border

            c = ws.cell(row=row, column=17, value=e.get('remark', ''))
            c.font = normal_font
            c.alignment = left
            c.border = border

            row += 1
            idx += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=f'{ym} 汇总：共 {month_total} 个事件，资料齐全 {month_complete} 个，待补 {month_total - month_complete} 个，完成率 {month_complete/month_total*100:.0f}%')
        c.font = header_font
        c.alignment = Alignment(horizontal='right', vertical='center')
        for col in range(1, 18):
            ws.cell(row=row, column=col).border = border
        row += 2

    widths = [6, 10, 8, 11, 11, 18, 22, 10, 9, 10, 10, 10, 10, 10, 26, 10, 15]
    for i, w in enumerate(widths, 1):
        col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = 'A4'

    wb.save(path)
    return path


def CostCalcWidget_num2chinese(n):
    if not n:
        return '零元整'
    digits = '零壹贰叁肆伍陆柒捌玖'
    units = ['', '拾', '佰', '仟', '万', '拾', '佰', '仟', '亿']
    integer_part = int(n)
    dec_part = round((n - integer_part) * 100)
    result = ''
    s = str(integer_part)
    for i, c in enumerate(reversed(s)):
        d = int(c)
        if d:
            result = digits[d] + units[i] + result
        else:
            if result and not result.startswith('零'):
                result = '零' + result
    result = result.rstrip('零') + '元'
    if dec_part == 0:
        result += '整'
    else:
        jiao = dec_part // 10
        fen = dec_part % 10
        if jiao:
            result += digits[jiao] + '角'
        if fen:
            result += digits[fen] + '分'
    return result
